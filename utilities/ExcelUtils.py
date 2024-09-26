import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file_path,sheetName):
    workbook= openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file_path,sheetName):
    workbook= openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file_path,sheetName,row,column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    return(sheet.cell(row,column).value)

def writeData(file_path,sheetName2,row,column,data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName2]
    sheet.cell(row,column).value=data
    workbook.save(file_path)
    return None

def fillRedColour(file_path,sheetName,row,column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(row,column).fill=redFill
    workbook.save(file_path)
    return  None

def fillGreenColour(file_path,sheetName,row,column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='00FF00',
                            end_color='00FF00',
                            fill_type='solid')
    sheet.cell(row,column).fill=greenFill
    workbook.save(file_path)
    return  None


